"""
build_icd.py — Xây lexicon ICD-10 từ ICD10.xlsx (song ngữ Anh–Việt).

Đầu ra: kb/icd_index.json  gồm:
  entries : [{code, en, vn, key_vn, key_en, level}]  (level 3 = mã 3 ký tự, 4 = leaf)
  exact   : {norm_key -> [codes]}     (khớp chính xác tên VN/EN)

Chạy: python kb/build_icd.py /path/to/ICD10.xlsx
"""
import json
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
from vntext import norm_key  # noqa

# Vị trí cột trong file (0-indexed), xác định bằng khảo sát dữ liệu:
COL_CODE = 17   # MÃ BỆNH  (vd A00.0)
COL_EN = 19     # DISEASE NAME WHO 2019 (ENGLISH)
COL_VN = 21     # TÊN BỆNH


def build(xlsx_path: str, out_path: str):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    entries = []
    exact = {}

    for r in ws.iter_rows(values_only=True):
        code = r[COL_CODE]
        if not code or not isinstance(code, str):
            continue
        code = code.strip()
        if len(code) < 3 or not code[0].isalpha():
            continue
        en = (r[COL_EN] or "").strip() if isinstance(r[COL_EN], str) else ""
        vn = (r[COL_VN] or "").strip() if isinstance(r[COL_VN], str) else ""
        level = 4 if "." in code else 3

        eid = len(entries)
        kv = norm_key(vn) if vn else ""
        ke = norm_key(en) if en else ""
        entries.append({"code": code, "en": en, "vn": vn,
                         "key_vn": kv, "key_en": ke, "level": level})

        for k in (kv, ke):
            if k:
                exact.setdefault(k, [])
                if code not in exact[k]:
                    exact[k].append(code)

    out = {"entries": entries, "exact": exact}
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False)
    print(f"[ICD] entries={len(entries)}  exact_keys={len(exact)}  -> {out_path}")


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/ICD10.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(__file__), "icd_index.json")
    build(xlsx, out)
